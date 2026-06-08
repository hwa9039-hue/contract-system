"""계약 엑셀 import 성공 시 DB 저장분을 역추출한 .xlsx 백업 저장 (NAS 백업 등).

환경변수 CONTRACT_IMPORT_EXCEL_BACKUP_DIR: 저장 디렉터리 (없거나 비었으면 백업 생략).
"""

from __future__ import annotations

import logging
import os
import re
from datetime import date, datetime, timezone
from decimal import Decimal, InvalidOperation
from pathlib import Path
from typing import Any

logger = logging.getLogger(__name__)

# openpyxl / Excel XML 에서 금지되는 제어 문자 (탭·LF·CR 제외)
_OPENPYXL_ILLEGAL_CHARACTERS_RE = re.compile(r"[\x00-\x08\x0b\x0c\x0e-\x1f]")


def _sanitize_openpyxl_text(value: Any) -> str:
    """주간보고 등 복붙 텍스트에 섞인 제어 문자 제거 — IllegalCharacterError 방지."""
    if value is None:
        return ""
    if not isinstance(value, str):
        value = str(value)
    return _OPENPYXL_ILLEGAL_CHARACTERS_RE.sub("", value)


# 프론트 handleExcelDownload / 엑셀 업로드 컬럼과 동일 순서·헤더
CONTRACT_BACKUP_HEADERS: tuple[tuple[str, str], ...] = (
    ("year", "사업년도"),
    ("segment", "구분"),
    ("refNo", "참고번호"),
    ("contractNo", "계약번호"),
    ("client", "발주처"),
    ("department", "담당부서"),
    ("contractMethod", "계약방식"),
    ("contractType", "계약분류"),
    ("identNo", "식별번호"),
    ("contractDate", "계약일자"),
    ("dueDate", "준공일자"),
    ("projectName", "사업명"),
    ("amount", "계약금액"),
    ("salesOwner", "영업담당자"),
    ("pm", "현장 PM"),
    ("note", "비고"),
)


def _resolved_backup_directory() -> Path | None:
    raw = (os.getenv("CONTRACT_IMPORT_EXCEL_BACKUP_DIR") or "").strip()
    if not raw:
        return None
    p = Path(raw).expanduser()
    try:
        p.mkdir(parents=True, exist_ok=True)
        return p.resolve()
    except OSError as exc:
        logger.warning("계약 엑셀 백업 디렉터리 생성 실패: %s (%s)", p, exc)
        return None


def _normalize_cell(value: Any) -> Any:
    if value is None:
        return ""
    if isinstance(value, datetime):
        return value.date().isoformat() if value.tzinfo is None else value.astimezone(timezone.utc).date().isoformat()
    if isinstance(value, date):
        return value.isoformat()
    if isinstance(value, bool):
        return str(value)
    if isinstance(value, (int, float)):
        return value
    return _sanitize_openpyxl_text(value)


def _format_amount_cell(value: Any) -> str:
    if value is None or value == "":
        return ""
    try:
        if isinstance(value, Decimal):
            n = int(value)
        elif isinstance(value, (int, float)):
            n = int(value)
        else:
            s = "".join(c for c in str(value) if c.isdigit() or c == "-").strip("-")
            n = int(Decimal(s)) if s else 0
        if n < 0:
            n = 0
        return f"{n:,}"
    except (InvalidOperation, ValueError, TypeError):
        return _sanitize_openpyxl_text(value)


def write_contract_import_excel_backup(
    *,
    inserted_contract_rows: list[dict[str, Any]],
    duplicate_items: list[str] | None = None,
) -> tuple[str | None, str | None]:
    """
    신규 insert된 계약 행만 시트 「계약현황」에 기록하고, 같은 파일에 「중복제외」 시트 추가.
    Returns (absolute_path, error_message). error_message가 None이면 성공.
    """
    dup = list(duplicate_items or [])
    if not inserted_contract_rows:
        return None, None

    target_dir = _resolved_backup_directory()
    if target_dir is None:
        # 환경변수 미설정 시 기능 생략(오류 아님)
        return None, None

    stamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    filename = f"계약현황_백업_{stamp}.xlsx"
    out_path = target_dir / filename

    try:
        from openpyxl import Workbook
        from openpyxl.styles import Alignment, Font

        wb = Workbook()
        ws_main = wb.active
        ws_main.title = "계약현황"
        hdr_font = Font(bold=True)
        for col, (_, label) in enumerate(CONTRACT_BACKUP_HEADERS, start=1):
            c = ws_main.cell(row=1, column=col, value=label)
            c.font = hdr_font

        row_idx = 2
        for item in inserted_contract_rows:
            for col, (key, _) in enumerate(CONTRACT_BACKUP_HEADERS, start=1):
                val = item.get(key)
                cell = ws_main.cell(row=row_idx, column=col)
                if key == "amount":
                    cell.value = _format_amount_cell(val)
                else:
                    cell.value = _normalize_cell(val)
                    cell.alignment = Alignment(vertical="top", wrap_text=True)
            row_idx += 1

        if dup:
            ws_dup = wb.create_sheet(title="중복제외")
            ws_dup.cell(row=1, column=1, value="중복·제외 항목 식별").font = hdr_font
            for i, label in enumerate(dup, start=2):
                ws_dup.cell(row=i, column=1, value=str(label))

        wb.save(out_path)

    except ImportError:
        logger.exception("openpyxl 미설치 — pip install openpyxl 필요")
        return None, "openpyxl 미설치로 엑셀 백업 불가"

    except OSError:
        logger.exception("계약 엑셀 백업 저장 실패: %s", out_path)
        return None, f"파일 저장 실패: {out_path}"

    except Exception:
        logger.exception("계약 엑셀 백업 작성 중 예외")
        return None, "알 수 없는 오류로 엑셀 백업 실패"

    logger.info("계약 import 엑셀 백업 저장: %s", out_path)
    return str(out_path), None
