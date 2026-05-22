"""전체 메뉴 DB 데이터를 화면 엑셀 다운로드와 동일한 형식의 .xlsx 로 내보냅니다.

NAS backup-postgres.sh 가 백업 세션마다 호출합니다.
환경변수 MENU_EXCEL_EXPORT_DIR: 출력 디렉터리 (필수).
"""

from __future__ import annotations

import logging
import os
import sys
from datetime import datetime, timezone
from pathlib import Path
from typing import Any, Callable

from app.contract_import_backup_xlsx import _format_amount_cell, _normalize_cell
from app.database import get_connection
from app.schemas import (
    row_to_calendar_manual_event,
    row_to_contract,
    row_to_document_register,
    row_to_excluded_project,
    row_to_install_case,
    row_to_materials_board_post,
    row_to_project_discovery,
    row_to_sales_register,
    row_to_weekly_work_report,
)

logger = logging.getLogger(__name__)

ExportColumn = tuple[str, str]
RowMapper = Callable[[dict], dict]
RowProjector = Callable[[dict], dict[str, Any]]


def _amount_display(value: Any) -> str:
    return _format_amount_cell(value)


def _project_contract(row: dict) -> dict[str, Any]:
    item = row_to_contract(row)
    return {
        "사업년도": _normalize_cell(item.get("year")),
        "구분": _normalize_cell(item.get("segment")),
        "참고번호": _normalize_cell(item.get("refNo")),
        "계약번호": _normalize_cell(item.get("contractNo")),
        "발주처": _normalize_cell(item.get("client")),
        "담당부서": _normalize_cell(item.get("department")),
        "계약방식": _normalize_cell(item.get("contractMethod")),
        "계약분류": _normalize_cell(item.get("contractType")),
        "식별번호": _normalize_cell(item.get("identNo")),
        "계약일자": _normalize_cell(item.get("contractDate")),
        "준공일자": _normalize_cell(item.get("dueDate")),
        "사업명": _normalize_cell(item.get("projectName")),
        "계약금액": _amount_display(item.get("amount")),
        "영업담당자": _normalize_cell(item.get("salesOwner")),
        "현장 PM": _normalize_cell(item.get("pm")),
        "비고": _normalize_cell(item.get("note")),
    }


def _project_sales(row: dict) -> dict[str, Any]:
    item = row_to_sales_register(row)
    return {
        "등록일": _normalize_cell(item.get("registerDate")),
        "발주처": _normalize_cell(item.get("client")),
        "프로젝트": _normalize_cell(item.get("projectName")),
        "사업금액": _amount_display(item.get("projectAmount")),
        "사업구분": _normalize_cell(item.get("projectCategory")),
        "담당자": _normalize_cell(item.get("manager")),
        "상태": _normalize_cell(item.get("projectStage")),
        "담당부서": _normalize_cell(item.get("department")),
        "세부내용": _normalize_cell(item.get("detail")),
        "출처": _normalize_cell(item.get("source")),
        "영업매칭": _normalize_cell(item.get("salesNote")),
        "영업 요청사항": _normalize_cell(item.get("actionRequest")),
    }


def _project_document(row: dict) -> dict[str, Any]:
    item = row_to_document_register(row)
    return {
        "등록일": _normalize_cell(item.get("docDate")),
        "문서번호": _normalize_cell(item.get("docNo")),
        "수신처 또는 발신처": _normalize_cell(item.get("senderReceiver")),
        "문서명 또는 제목": _normalize_cell(item.get("title")),
        "접수 또는 발송형태": _normalize_cell(item.get("method")),
        "수신자 또는 작성자": _normalize_cell(item.get("writer")),
        "비고": _normalize_cell(item.get("note")),
    }


def _project_discovery(row: dict) -> dict[str, Any]:
    item = row_to_project_discovery(row)
    return {
        "건축정보일자": _normalize_cell(item.get("permitDate")),
        "확인": _normalize_cell(item.get("checkStatus")),
        "영업자": _normalize_cell(item.get("salesTarget")),
        "사업구분": _normalize_cell(item.get("projectCategory")),
        "발주처": _normalize_cell(item.get("client")),
        "사업명": _normalize_cell(item.get("projectName")),
        "사업금액": _amount_display(item.get("projectAmount")),
        "준공시기": _normalize_cell(item.get("completionPeriod")),
        "담당자": _normalize_cell(item.get("manager")),
        "비고": _normalize_cell(item.get("note")),
    }


def _project_excluded(row: dict) -> dict[str, Any]:
    item = row_to_excluded_project(row)
    return {
        "등록일": _normalize_cell(item.get("writeDate")),
        "공개일": _normalize_cell(item.get("openDate")),
        "상태": _normalize_cell(item.get("category")),
        "검색어": _normalize_cell(item.get("keyword")),
        "작성자": _normalize_cell(item.get("writer")),
        "사업명": _normalize_cell(item.get("projectName")),
        "발주처": _normalize_cell(item.get("client")),
        "사업금액": _amount_display(item.get("projectAmount")),
        "제외 사유": _normalize_cell(item.get("exclusionReason")),
    }


def _project_work_report(row: dict) -> dict[str, Any]:
    item = row_to_weekly_work_report(row)
    return {
        "보고일": _normalize_cell(item.get("reportDate") or item.get("date")),
        "주차": _normalize_cell(item.get("weekNumber")),
        "담당자": _normalize_cell(item.get("assignee") or item.get("user")),
        "팀": _normalize_cell(item.get("team")),
        "구분": _normalize_cell(item.get("category") or item.get("section")),
        "내용": _normalize_cell(item.get("content")),
    }


def _project_calendar(row: dict) -> dict[str, Any]:
    item = row_to_calendar_manual_event(row)
    return {
        "시작일": _normalize_cell(item.get("dateStart")),
        "종료일": _normalize_cell(item.get("dateEnd")),
        "일정 내용": _normalize_cell(item.get("title")),
        "영업담당자": _normalize_cell(item.get("owner")),
        "현장 PM": _normalize_cell(item.get("pm")),
        "비고": _normalize_cell(item.get("note")),
    }


def _project_install_case(row: dict) -> dict[str, Any]:
    item = row_to_install_case(row)
    specs = item.get("specs") if isinstance(item.get("specs"), dict) else {}
    return {
        "사업명": _normalize_cell(item.get("projectName")),
        "사업년도": _normalize_cell(item.get("year")),
        "대분류": _normalize_cell(item.get("environment")),
        "중분류": _normalize_cell(item.get("middleCategory")),
        "소분류": _normalize_cell(item.get("audience")),
        "용도": _normalize_cell(item.get("purpose")),
        "발주처": _normalize_cell(item.get("client")),
        "표출부 사이즈": _normalize_cell(specs.get("displayArea")),
        "LED Pitch": _normalize_cell(specs.get("ledPitch")),
        "MODULE 크기": _normalize_cell(specs.get("moduleSize")),
        "MODULE 수량": _normalize_cell(specs.get("moduleQty")),
        "해상도": _normalize_cell(specs.get("resolution")),
        "설치유형": _normalize_cell(specs.get("installType")),
    }


def _project_materials_board(row: dict) -> dict[str, Any]:
    item = row_to_materials_board_post(row)
    files = item.get("files") if isinstance(item.get("files"), list) else []
    names = ", ".join(
        str(f.get("name") or "").strip()
        for f in files
        if isinstance(f, dict) and str(f.get("name") or "").strip()
    )
    return {
        "제목": _normalize_cell(item.get("title")),
        "등록일": _normalize_cell(item.get("registeredAt")),
        "다운로드 수": _normalize_cell(item.get("downloadCount")),
        "첨부파일": names,
        "내용": _normalize_cell(item.get("content")),
    }


MENU_EXPORT_SPECS: tuple[dict[str, Any], ...] = (
    {
        "file_prefix": "계약현황",
        "sheet_title": "계약현황",
        "table": "contracts_rows",
        "order_by": '"contractDate" desc nulls last, year desc nulls last',
        "project": _project_contract,
    },
    {
        "file_prefix": "영업관리대장",
        "sheet_title": "영업관리대장",
        "table": "sales_register_rows",
        "order_by": '"registerDate" desc nulls last, "createdAt" desc nulls last',
        "project": _project_sales,
    },
    {
        "file_prefix": "문서수발신대장",
        "sheet_title": "문서수발신대장",
        "table": "document_register_rows",
        "order_by": '"docDate" desc nulls last, "createdAt" desc nulls last',
        "project": _project_document,
    },
    {
        "file_prefix": "건축정보",
        "sheet_title": "건축정보",
        "table": "project_discovery_rows",
        "order_by": '"permitDate" desc nulls last, "createdAt" desc nulls last',
        "project": _project_discovery,
    },
    {
        "file_prefix": "사업검색이력",
        "sheet_title": "사업검색이력",
        "table": "excluded_projects_rows",
        "order_by": '"writeDate" desc nulls last, "createdAt" desc nulls last',
        "project": _project_excluded,
    },
    {
        "file_prefix": "주간업무보고서",
        "sheet_title": "주간업무보고서",
        "table": "weekly_work_reports_rows",
        "order_by": '"date" desc nulls last, "reportDate" desc nulls last, order_index asc nulls last',
        "project": _project_work_report,
    },
    {
        "file_prefix": "캘린더_기타일정",
        "sheet_title": "캘린더기타",
        "table": "calendar_manual_events",
        "order_by": '"dateStart" desc nulls last, "createdAt" desc nulls last',
        "project": _project_calendar,
        "optional_table": True,
    },
    {
        "file_prefix": "설치사례",
        "sheet_title": "설치사례",
        "table": "install_cases_rows",
        "order_by": '"createdAt" desc nulls last',
        "project": _project_install_case,
    },
    {
        "file_prefix": "게시판",
        "sheet_title": "게시판",
        "table": "materials_board_posts",
        "order_by": '"registeredAt" desc nulls last, "createdAt" desc nulls last',
        "project": _project_materials_board,
    },
)


def _table_exists(cursor, table: str) -> bool:
    cursor.execute(
        """
        select 1
        from information_schema.tables
        where table_schema = 'public' and table_name = %s
        """,
        (table,),
    )
    return cursor.fetchone() is not None


def _fetch_table_rows(cursor, table: str, order_by: str) -> list[dict]:
    cursor.execute(f'select * from "{table}" order by {order_by}')
    return list(cursor.fetchall())


def _write_workbook(path: Path, sheet_title: str, rows: list[dict[str, Any]]) -> None:
    from openpyxl import Workbook
    from openpyxl.styles import Font

    wb = Workbook()
    ws = wb.active
    ws.title = sheet_title[:31]

    headers: list[str] = []
    if rows:
        headers = list(rows[0].keys())
        hdr_font = Font(bold=True)
        for col, label in enumerate(headers, start=1):
            cell = ws.cell(row=1, column=col, value=label)
            cell.font = hdr_font
        for row_idx, row in enumerate(rows, start=2):
            for col_idx, key in enumerate(headers, start=1):
                ws.cell(row=row_idx, column=col_idx, value=row.get(key, ""))
    else:
        ws.cell(row=1, column=1, value="(데이터 없음)")

    path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(path)


def export_all_menu_excel_backups(output_dir: Path, stamp: str | None = None) -> list[Path]:
    stamp = stamp or datetime.now().strftime("%Y%m%d_%H%M%S")
    output_dir.mkdir(parents=True, exist_ok=True)
    written: list[Path] = []

    with get_connection() as connection:
        with connection.cursor() as cursor:
            for spec in MENU_EXPORT_SPECS:
                table = spec["table"]
                if spec.get("optional_table") and not _table_exists(cursor, table):
                    logger.info("skip excel export (table missing): %s", table)
                    continue
                if not _table_exists(cursor, table):
                    raise RuntimeError(f"Required table missing for excel export: {table}")

                db_rows = _fetch_table_rows(cursor, table, spec["order_by"])
                projected = [spec["project"](row) for row in db_rows]
                out_path = output_dir / f"{spec['file_prefix']}_백업_{stamp}.xlsx"
                _write_workbook(out_path, spec["sheet_title"], projected)
                written.append(out_path)
                logger.info("excel backup: %s (%s rows)", out_path.name, len(projected))

    return written


def resolved_export_directory() -> Path | None:
    raw = (os.getenv("MENU_EXCEL_EXPORT_DIR") or "").strip()
    if not raw:
        return None
    return Path(raw).expanduser().resolve()


def main() -> int:
    logging.basicConfig(level=logging.INFO, format="%(message)s")
    target = resolved_export_directory()
    if target is None:
        print("ERROR: MENU_EXCEL_EXPORT_DIR is not set.", file=sys.stderr)
        return 1
    try:
        files = export_all_menu_excel_backups(target)
    except Exception:
        logger.exception("menu excel export failed")
        return 1
    print(f"Excel export OK: {target} ({len(files)} files)")
    for path in files:
        print(f"  - {path.name}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
